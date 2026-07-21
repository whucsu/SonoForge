"""Import reference data from Excel files."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


def import_excel_file(path: Path) -> dict[str, Any]:
    """Import reference data from .xlsx/.xls file.

    Expected format:
    - Sheet per topic (name = topic name)
    - Row 1: headers (name, unit, norm_male_low, norm_male_high, ...)
    - Row 2+: parameter data

    Returns dict compatible with ReferenceModel.from_dict().
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    topics = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        # Parse headers
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value or "").strip().lower())

        # Parse rows
        parameters = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            param: dict[str, Any] = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    field = headers[col_idx]
                    if field in ("id", "name", "unit", "pathology_desc", "source"):
                        param[field] = str(value or "")
                    elif field in ("norm_male_low", "norm_male_high"):
                        if "norm_male" not in param:
                            param["norm_male"] = {}
                        param["norm_male"][field.replace("norm_male_", "")] = _parse_num(value)
                    elif field in ("norm_female_low", "norm_female_high"):
                        if "norm_female" not in param:
                            param["norm_female"] = {}
                        param["norm_female"][field.replace("norm_female_", "")] = _parse_num(value)

            if param.get("id"):
                parameters.append(param)

        if parameters:
            slug = sheet_name.lower().replace(" ", "_")
            topics.append(
                {
                    "name": sheet_name,
                    "slug": slug,
                    "pathologies": [
                        {
                            "name": sheet_name,
                            "slug": f"{slug}_all",
                            "parameters": parameters,
                        }
                    ],
                }
            )

    return {"topics": topics}


def _parse_num(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None
